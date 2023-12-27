import io
import pycurl
from pathlib import Path
import pandas as pd
from pandas import Timestamp as ts, Timedelta as td
from openpyxl import load_workbook

from caiso_logging import DataLogger

class CurtailmentDownloader:
    '''
    A class to manage downloads of CAISO daily curtailment reports.
    '''
    start_date = ts(2021,6,18)
    def __init__(self,download_directory_path:Path,log_path:Path):
        log_dtypes = {
            'effective_date' : 'datetime64[D]',
            'download_path' : 'string',
        }
        self.logger = DataLogger(dtypes=log_dtypes,log_path=log_path,delimiter=',')
        self.download_directory_path = download_directory_path

    def url_by_date(self,date:ts):
        '''
        Generates the url to a CAISO prior trade day curtailment report for the
        given date

        Parameters:
            date - a Pandas Timestamp object representing a given day

        Returns:
            A string url pointing to a report if it exists for the given day.
        '''
        return date.strftime('http://www.caiso.com/Documents/Curtailed-non-operational-generator-prior-trade-date-report-%Y%m%d.xlsx')

    def path_by_date(self,date:ts):
        '''
        Generates the path to which a report is or will be saved for the given
        date.

        Parameters:
            date - a Pandas Timestamp object representing a given day

        Returns:
            A pathlib Path object pointing to a local location for a report to
            be saved.
        '''
        return self.download_directory_path / date.strftime('PriorTradeDateCurtailments_%Y-%m-%d.xlsx')

    def download_report(self,date):
        '''
        Downloads a prior trade day curtailments report from the CAISO website,
        if available.

        Parameters:
            date - a Pandas Timestamp object representing a given day

        Side Effects:
            Downloads and saves an Excel spreadsheet file.
            Prints actions to console

        Returns:
            Integer representing status of download:
                -1 if unable to download file (response from CAISO website != 200)
                0 if file is already downloaded according to log
                1 if file successfully downloaded and saved to location
        '''
        if (date==self.logger.data.loc[:,'effective_date']).any():
            download_date = self.logger.data.loc[(self.logger.data.loc[:,'effective_date']==date),'log_timestamp'].iloc[0]
            print('Skipping Download for {} [Already downloaded {}]'.format(date.strftime('%Y-%m-%d'),download_date.strftime('%Y-%m-%d %H:%M:%S')))
            return 0
        else:
            url = self.url_by_date(date)
            download_path = self.path_by_date(date)
            with download_path.open('wb') as f:
                c = pycurl.Curl()
                c.setopt(c.URL,url)
                c.setopt(c.WRITEDATA,f)
                c.perform()
                c.close()
                self.logger.log(pd.Series({
                    'effective_date' : date,
                    'download_path' : download_path,
                }))
                self.logger.commit()
            print('Downloading for {}'.format(date.strftime('%Y-%m-%d')))
            return c.RESPONSE_CODE

    def download_all(self):
        '''
        Downloads all prior trade day curtailments report from the CAISO website,
        from the initial date available (June 18, 2021) to yesterday.

        Parameters:
            None

        Side Effects:
            Calls download_report method.

        Returns:
            None
        '''
        today = ts.now().replace(hour=0,minute=0,second=0,microsecond=0)
        date_range = [ts.fromordinal(d) for d in range(self.start_date.toordinal(),today.toordinal())]
        for date in date_range:
            self.download_report(date)

    def extract_by_nature_of_work(self,nature_of_work:str):
        return self.extract_by_columns([('NATURE OF WORK',nature_of_work)])

    def extract_by_columns(self,kvps:list,effective_dates:list=None):
        '''
        Extracts rows from all downloaded reports filtered by a set of key-
        value pairs in the input kvps list. Rows must match all pairs in order
        to be included, and the results are returned as a Pandas dataframe.

        Parameters:
            kvps - a list containing key-value pairs as tuples with the first
                element being a string matching the label of one of the columns in the
                reports and the second element a value in the matching column
            effective_dates - a list containing datetime objects representing
                dates corresponding to the effective_date column of the log
                associated with filenames from which to extract data

        Returns:
            Dataframe containing rows matching input key-value pairs
        '''
        column_names = [
            'OUTAGE MRID',
            'RESOURCE NAME',
            'RESOURCE ID',
            'OUTAGE TYPE',
            'NATURE OF WORK',
            'CURTAILMENT START DATE TIME',
            'CURTAILMENT END DATE TIME',
            'CURTAILMENT MW',
            'RESOURCE PMAX MW',
            'NET QUALIFYING CAPACITY MW',
        ]
        df = pd.DataFrame(columns=column_names)
        if effective_dates is None:
            download_path_strs = list(self.logger.data.loc[:,'download_path'])
        else:
            download_path_strs = list(self.logger.data.loc[self.logger.data.loc[:,'effective_date'].apply(lambda d: d in effective_dates),'download_path'])
        for download_path_str in download_path_strs:
            with Path(download_path_str).open('rb') as f:
                print('Reading '+Path(download_path_str).name)
                in_mem_file = io.BytesIO(f.read())
                wb = load_workbook(in_mem_file,data_only=True,read_only=True)
            ws = wb['PREV_DAY_OUTAGES']
            new_data = {k:[] for k in column_names}
            for data_range_row in ws.iter_rows(min_row=11):
                if len(data_range_row)>0:
                    data_values= [data_range_row[c].value for c in [1,2,4,5,6,7,8,9,10,12]]
                    #data_values= [c.value for c in filter(lambda c:c.value is not None,data_range_row)]
                    for column_name,data_value in zip(column_names,data_values):
                        new_data[column_name].append(data_value)
            new_dataframe = pd.DataFrame(new_data)
            filter_keys = pd.Series([True]*len(new_dataframe))
            for kvp in kvps:
                filter_keys &= (new_dataframe.loc[:,kvp[0]]==kvp[1])
            df = df.append(new_dataframe.loc[filter_keys,:],ignore_index=True)
        return df

    def extract_all(self,effective_dates:list=[]):
        '''
        Extracts data from all downloaded reports without filtering.

        Parameters:
            effective_dates - a list containing datetime objects representing
                dates corresponding to the effective_date column of the log
                associated with filenames from which to extract data

        Returns:
            Dataframe containing data from curtailment reports matching
            the given effective dates.
        '''
        column_names = [
            'OUTAGE MRID',
            'RESOURCE NAME',
            'RESOURCE ID',
            'OUTAGE TYPE',
            'NATURE OF WORK',
            'CURTAILMENT START DATE TIME',
            'CURTAILMENT END DATE TIME',
            'CURTAILMENT MW',
            'RESOURCE PMAX MW',
            'NET QUALIFYING CAPACITY MW',
            'OUTAGE STATUS',
            'RES TYPE',
            'MKTORGANIZATION MRID',
            'BAA'
        ]
        df = pd.DataFrame(columns=column_names)
        if len(effective_dates)>0:
            download_path_strs = list(self.logger.data.loc[self.logger.data.loc[:,'effective_date'].apply(lambda d: d in effective_dates),'download_path'])
        else:
            download_path_strs = list(self.logger.data.loc[:,'download_path'])
        for effective_date,download_path_str in zip(effective_dates,download_path_strs):
            with Path(download_path_str).open('rb') as f:
                print('Reading '+Path(download_path_str).name)
                in_mem_file = io.BytesIO(f.read())
                wb = load_workbook(in_mem_file,data_only=True,read_only=True)
            ws = wb['PREV_DAY_OUTAGES']
            new_data = {k:[] for k in column_names}
            # find header row:
            header_row_number = 1
            while True:
                header_row = list(map(lambda x:x.value,ws[header_row_number]))
                if column_names[0] in header_row or header_row_number>100:
                    break
                else:
                    header_row_number += 1

            columns = {k: header_row.index(k) if k in header_row else None for k in column_names}
            if header_row_number<100:
                for data_range_row in ws.iter_rows(min_row=header_row_number+1):
                    if len(data_range_row)>0:
                        for column_name,column_number in columns.items():
                            if column_number is not None:
                                new_data[column_name].append(data_range_row[column_number].value)
                            else:
                                new_data[column_name].append(None)
                new_dataframe = pd.DataFrame(new_data)
                # Constrain curtailment hours within trade day:
                new_dataframe.loc[:,'CURTAILMENT START DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT START DATE TIME'].apply(lambda t: max(t,effective_date.replace(hour=0,minute=0,second=0)))
                new_dataframe.loc[:,'CURTAILMENT END DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT END DATE TIME'].fillna(effective_date+td(days=1))
                new_dataframe.loc[:,'CURTAILMENT END DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT END DATE TIME'].apply(lambda t: min(t,effective_date.replace(hour=23,minute=59,second=59)))
                df = df.append(new_dataframe,ignore_index=True)
            else:
                pass
        return df

    def calculate_monthly_outage_rates(self,resource_ids:list,effective_month:ts):
        start_of_month = effective_month.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
        end_of_month = effective_month.replace(year=effective_month.year+int(effective_month.month/12),month=(effective_month.month)%12+1,day=1) + td(microseconds=-1)
        effective_dates = [ts(effective_month.year,effective_month.month,d) for d in range(1,end_of_month.day+1)]
        month_duration = end_of_month.replace(hour=23,minute=59,second=59,microsecond=999999) - effective_month.replace(day=1,hour=0,minute=0,second=0,microsecond=0)

        # read curtailment reports for effective month:
        forced_outages = self.extract_by_columns([['OUTAGE TYPE','FORCED']],effective_dates=effective_dates)

        # remove any curtailments without start or end times:
        forced_outages = forced_outages.dropna(axis='index',how='any',subset=['CURTAILMENT START DATE TIME','CURTAILMENT END DATE TIME'])

        # truncate curtailments extending before or after effective month:
        forced_outages.loc[:,'CURTAILMENT START DATE TIME'] = forced_outages.loc[:,'CURTAILMENT START DATE TIME'].map(lambda t:max(t,start_of_month))
        forced_outages.loc[:,'CURTAILMENT END DATE TIME'] = forced_outages.loc[:,'CURTAILMENT END DATE TIME'].map(lambda t:min(t,end_of_month))

        # use only last curtailment report for given mrid and start time:
        forced_outages = forced_outages.groupby(['OUTAGE MRID','CURTAILMENT START DATE TIME']).last().reset_index()
        # alternate: use only first curtailment report for given mrid and start time:
        # forced_outages = forced_outages.groupby(['OUTAGE MRID','CURTAILMENT START DATE TIME']).first().reset_index()

        # calculate outage durations:
        forced_outages.loc[:,'CURTAILMENT DURATION'] = forced_outages.loc[:,'CURTAILMENT END DATE TIME'] - forced_outages.loc[:,'CURTAILMENT START DATE TIME']
        forced_outages.loc[:,'OUTAGE MWH'] = forced_outages.loc[:,'CURTAILMENT MW'] * forced_outages.loc[:,'CURTAILMENT DURATION'].map(lambda x:x.total_seconds()/3600)
        outage_rates = pd.DataFrame({'RESOURCE ID':resource_ids}).set_index('RESOURCE ID')
        outage_rates = outage_rates.merge(forced_outages.loc[:,['RESOURCE ID','RESOURCE PMAX MW']].groupby('RESOURCE ID').mean(),left_index=True,right_index=True,how='left')
        outage_rates = outage_rates.rename(columns={'RESOURCE PMAX MW':'MW CAPACITY'})
        outage_rates = outage_rates.merge(forced_outages.loc[:,['RESOURCE ID','CURTAILMENT DURATION']].groupby('RESOURCE ID').sum().rename(columns={'CURTAILMENT DURATION':'TOTAL OUTAGE TIME'}),left_index=True,right_index=True)
        outage_rates = outage_rates.merge(forced_outages.loc[:,['RESOURCE ID','CURTAILMENT MW']].groupby('RESOURCE ID').sum().rename(columns={'CURTAILMENT MW':'SUM CURTAILMENT MW'}),left_index=True,right_index=True)
        outage_rates = outage_rates.merge(forced_outages.loc[:,['RESOURCE ID','OUTAGE MWH']].groupby('RESOURCE ID').sum(),left_index=True,right_index=True)
        outage_rates.loc[:,'TIME-WEIGHTED AVERAGE MW CURTAILMENT'] = outage_rates.loc[:,'OUTAGE MWH'] / outage_rates.loc[:,'TOTAL OUTAGE TIME'].map(lambda x:x.total_seconds()/3600)
        outage_rates.loc[:,'FORCED OUTAGE RATE BY TIME'] = outage_rates.loc[:,'TOTAL OUTAGE TIME'] / month_duration
        outage_rates.loc[:,'FORCED OUTAGE RATE BY MWH'] = outage_rates.loc[:,'OUTAGE MWH'] / (outage_rates.loc[:,'MW CAPACITY'] * month_duration.total_seconds() / 3600)
        outage_rates.loc[:,'MONTH'] = start_of_month

        # return outages matching input list; missing resources are assigned 0% outage rate
        return outage_rates.loc[:,['MONTH','MW CAPACITY','TOTAL OUTAGE TIME','SUM CURTAILMENT MW','OUTAGE MWH','TIME-WEIGHTED AVERAGE MW CURTAILMENT','FORCED OUTAGE RATE BY TIME','FORCED OUTAGE RATE BY MWH']].reset_index()

if __name__=='__main__':
    start_time = ts.now()
    first_date = ts(2021,6,18)
    last_date = ts(start_time.date())
    date_range = [first_date + td(days=d) for d in range((last_date-first_date).days)]
    curtailment_downloader = CurtailmentDownloader(
        download_directory_path=Path(r'M:\Users\RH2\src\caiso_curtailments\caiso_curtailment_reports'),
        log_path= Path(r'M:\Users\RH2\src\caiso_curtailments\caiso_curtailment_reports\download_log.csv')
    )
    curtailment_downloader.download_all()

    # new_dates = list(curtailment_downloader.logger.data.loc[(curtailment_downloader.logger.data.loc[:,'log_timestamp']>start_time),'effective_date'])

    new_dates = list(
        filter(
            lambda d: d not in list(dict.fromkeys(curtailment_downloader.logger.data.loc[:,'effective_date'])),
            date_range
        )
    )

    if Path('M:\\Users\RH2\\src\\caiso_curtailments\\results\\curtailments_all.csv').is_file():
        df0 = pd.read_csv(Path('M:\\Users\RH2\\src\\caiso_curtailments\\results\\curtailments_all.csv'))
    else:
        df0 = pd.DataFrame(
            columns=[
                'OUTAGE MRID',
                'RESOURCE NAME',
                'RESOURCE ID',
                'OUTAGE TYPE',
                'NATURE OF WORK',
                'CURTAILMENT START DATE TIME',
                'CURTAILMENT END DATE TIME',
                'CURTAILMENT MW',
                'RESOURCE PMAX MW',
                'NET QUALIFYING CAPACITY MW',
                'OUTAGE STATUS',
                'RES TYPE',
                'MKTORGANIZATION MRID',
                'BAA'
            ]
        )
    # df1 = curtailment_downloader.extract_by_columns([('NATURE OF WORK','AMBIENT_DUE_TO_TEMP')],effective_dates=new_dates,impute_zeros=False)
    df1 = curtailment_downloader.extract_all(effective_dates=date_range)
    df0 = df0.append(df1,ignore_index=True)
    # df0.to_csv(Path('M:\\Users\\RH2\\src\\caiso_curtailments\\results\\curtailments_ambient_due_to_temp.csv'),index=False)
    df0.to_csv(Path('M:\\Users\\RH2\\src\\caiso_curtailments\\results\\curtailments_all.csv'),index=False)
    resource_ids = pd.DataFrame(df0.loc[:,'RESOURCE ID'].unique(),columns=['RESOURCE ID'])
    resource_ids.to_csv(Path('M:\\Users\\RH2\\src\\caiso_curtailments\\geospatial\\curtailed_resources.csv'),index=False)